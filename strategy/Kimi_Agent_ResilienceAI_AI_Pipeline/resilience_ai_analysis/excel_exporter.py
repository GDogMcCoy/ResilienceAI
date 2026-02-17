"""
ResilienceAI - Excel Export Module
Export vulnerability data as formatted Excel workbooks
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
from config import PROCESSED_DIR, REPORTS_DIR

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """Export county vulnerability data as formatted Excel workbooks"""
    
    # Color scheme for risk levels
    RISK_COLORS = {
        'High': '#DC3545',      # Red
        'Medium': '#FFC107',    # Yellow
        'Low': '#28A745',       # Green
        'Critical': '#721C24',  # Dark red
        'Unknown': '#6C757D'    # Gray
    }
    
    # Header styling
    HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    
    def __init__(self, df=None):
        """Initialize exporter with county data"""
        if df is None:
            path = PROCESSED_DIR / "county_features.csv"
            if path.exists():
                self.df = pd.read_csv(path, dtype={"fips": str})
            else:
                self.df = None
        else:
            self.df = df
    
    def export_workbook(
        self,
        sheets: List[str] = None,
        output_path: str = None,
        filters: Dict[str, Any] = None,
        include_charts: bool = True
    ) -> Dict:
        """
        Export data as multi-sheet Excel workbook
        
        Args:
            sheets: List of sheets to include
            output_path: Output file path
            filters: Filters to apply to data
            include_charts: Include charts in workbook
            
        Returns:
            Dictionary with export metadata
        """
        if self.df is None:
            return {"error": "Data not loaded"}
        
        if not HAS_OPENPYXL:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}
        
        # Apply filters
        df = self.df.copy()
        if filters:
            df = self._apply_filters(df, filters)
        
        # Default sheets
        if sheets is None:
            sheets = ['Summary', 'County Data', 'Risk Analysis', 'Infrastructure']
        
        # Create output path
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = REPORTS_DIR / f"resilienceai_export_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate sheets
        sheet_generators = {
            'Summary': self._create_summary_sheet,
            'County Data': self._create_county_data_sheet,
            'Risk Analysis': self._create_risk_analysis_sheet,
            'Infrastructure': self._create_infrastructure_sheet,
            'Demographics': self._create_demographics_sheet,
            'Disaster History': self._create_disaster_sheet,
            'High Risk Counties': self._create_high_risk_sheet,
            'State Summary': self._create_state_summary_sheet
        }
        
        for sheet_name in sheets:
            if sheet_name in sheet_generators:
                ws = wb.create_sheet(title=sheet_name)
                sheet_generators[sheet_name](ws, df, include_charts)
        
        # Save workbook
        wb.save(output_path)
        
        return {
            "output_path": str(output_path),
            "sheets": sheets,
            "row_count": len(df),
            "filters_applied": filters
        }
    
    def _apply_filters(self, df: pd.DataFrame, filters: Dict) -> pd.DataFrame:
        """Apply filters to dataframe"""
        result = df.copy()
        
        if 'state' in filters:
            state = filters['state']
            result = result[result['county_name'].str.contains(f', {state}$', regex=True, na=False)]
        
        if 'risk_level' in filters:
            result = result[result['risk_level'] == filters['risk_level']]
        
        if 'min_risk_score' in filters:
            result = result[result['risk_score'] >= filters['min_risk_score']]
        
        if 'max_risk_score' in filters:
            result = result[result['risk_score'] <= filters['max_risk_score']]
        
        if 'compound_risk' in filters and filters['compound_risk']:
            result = result[result['compound_risk_flag'] == True]
        
        if 'zero_redundancy' in filters and filters['zero_redundancy']:
            result = result[result['zero_redundancy_flag'] == True]
        
        return result
    
    def _create_summary_sheet(self, ws, df, include_charts):
        """Create executive summary sheet"""
        # Title
        ws['A1'] = 'ResilienceAI Export Summary'
        ws['A1'].font = Font(bold=True, size=16, color='1F4E78')
        ws.merge_cells('A1:D1')
        
        # Metadata
        ws['A3'] = 'Generated:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Total Counties:'
        ws['B4'] = len(df)
        ws['A5'] = 'States:'
        ws['B5'] = df['county_name'].str.extract(r', ([\w ]+)$')[0].nunique()
        
        # Risk distribution
        ws['A7'] = 'Risk Distribution'
        ws['A7'].font = Font(bold=True, size=12)
        
        risk_dist = df['risk_level'].value_counts()
        ws['A8'] = 'Risk Level'
        ws['B8'] = 'Count'
        ws['C8'] = 'Percentage'
        
        for idx, (level, count) in enumerate(risk_dist.items(), start=9):
            ws[f'A{idx}'] = level
            ws[f'B{idx}'] = count
            ws[f'C{idx}'] = f"{(count / len(df) * 100):.1f}%"
            
            # Color code by risk level
            if level in self.RISK_COLORS:
                ws[f'A{idx}'].fill = PatternFill(
                    start_color=self.RISK_COLORS[level],
                    end_color=self.RISK_COLORS[level],
                    fill_type='solid'
                )
        
        # Key metrics
        ws['A14'] = 'Key Metrics'
        ws['A14'].font = Font(bold=True, size=12)
        
        metrics = [
            ('Average Risk Score', df['risk_score'].mean()),
            ('High Risk Counties', (df['risk_score'] >= 0.7).sum()),
            ('Compound Risk Counties', df['compound_risk_flag'].sum() if 'compound_risk_flag' in df.columns else 0),
            ('Zero Redundancy Counties', df['zero_redundancy_flag'].sum() if 'zero_redundancy_flag' in df.columns else 0),
            ('Average Vulnerability Index', df['vulnerability_index'].mean()),
            ('Average Isolation Index', df['isolation_index'].mean())
        ]
        
        for idx, (metric, value) in enumerate(metrics, start=15):
            ws[f'A{idx}'] = metric
            ws[f'B{idx}'] = value
            if isinstance(value, float):
                ws[f'B{idx}'].number_format = '0.000'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_county_data_sheet(self, ws, df, include_charts):
        """Create detailed county data sheet"""
        # Define columns
        columns = [
            'fips', 'county_name', 'risk_score', 'risk_level',
            'vulnerability_index', 'isolation_index', 'total_population',
            'poverty_pct', 'elderly_pct', 'disability_pct', 'uninsured_pct',
            'disaster_count', 'disaster_count_recent',
            'dist_nearest_hospitals_km', 'dist_nearest_fire_stations_km',
            'count_hospitals_50km', 'count_fire_stations_50km',
            'compound_risk_flag', 'zero_redundancy_flag'
        ]
        
        # Filter to available columns
        available_cols = [c for c in columns if c in df.columns]
        
        # Write headers
        for col_idx, col_name in enumerate(available_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, col_name in enumerate(available_cols, 1):
                value = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if pd.isna(value):
                    cell.value = ''
                elif col_name in ['compound_risk_flag', 'zero_redundancy_flag']:
                    cell.value = 'Yes' if value else 'No'
                else:
                    cell.value = value
                
                # Format numbers
                if col_name.endswith('_pct') or col_name == 'risk_score':
                    cell.number_format = '0.00'
                elif col_name.endswith('_index'):
                    cell.number_format = '0.000'
                
                # Color code risk level
                if col_name == 'risk_level' and value in self.RISK_COLORS:
                    cell.fill = PatternFill(
                        start_color=self.RISK_COLORS[value],
                        end_color=self.RISK_COLORS[value],
                        fill_type='solid'
                    )
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(available_cols, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(len(col_name) + 2, 12)
    
    def _create_risk_analysis_sheet(self, ws, df, include_charts):
        """Create risk analysis sheet with charts"""
        # Risk score distribution
        ws['A1'] = 'Risk Score Distribution'
        ws['A1'].font = Font(bold=True, size=12)
        
        bins = [0, 0.3, 0.5, 0.7, 0.85, 1.0]
        labels = ['Very Low (0-0.3)', 'Low (0.3-0.5)', 'Medium (0.5-0.7)', 'High (0.7-0.85)', 'Critical (0.85-1.0)']
        
        df['risk_bin'] = pd.cut(df['risk_score'], bins=bins, labels=labels, include_lowest=True)
        risk_dist = df['risk_bin'].value_counts().sort_index()
        
        ws['A3'] = 'Risk Range'
        ws['B3'] = 'Count'
        ws['C3'] = 'Percentage'
        
        for idx, (label, count) in enumerate(risk_dist.items(), start=4):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = count
            ws[f'C{idx}'] = count / len(df)
            ws[f'C{idx}'].number_format = '0.0%'
        
        # Add chart
        if include_charts and HAS_OPENPYXL:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Risk Score Distribution"
            chart.y_axis.title = 'Count'
            
            data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(risk_dist))
            cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(risk_dist))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "E3")
        
        # Top 20 highest risk counties
        ws['A11'] = 'Top 20 Highest Risk Counties'
        ws['A11'].font = Font(bold=True, size=12)
        
        top_risk = df.nlargest(20, 'risk_score')[['county_name', 'risk_score', 'risk_level', 'vulnerability_index', 'isolation_index']]
        
        ws['A13'] = 'County'
        ws['B13'] = 'Risk Score'
        ws['C13'] = 'Risk Level'
        ws['D13'] = 'Vulnerability'
        ws['E13'] = 'Isolation'
        
        for idx, (_, row) in enumerate(top_risk.iterrows(), start=14):
            ws[f'A{idx}'] = row['county_name']
            ws[f'B{idx}'] = row['risk_score']
            ws[f'B{idx}'].number_format = '0.000'
            ws[f'C{idx}'] = row['risk_level']
            ws[f'D{idx}'] = row['vulnerability_index']
            ws[f'D{idx}'].number_format = '0.000'
            ws[f'E{idx}'] = row['isolation_index']
            ws[f'E{idx}'].number_format = '0.000'
            
            # Color code
            if row['risk_level'] in self.RISK_COLORS:
                ws[f'C{idx}'].fill = PatternFill(
                    start_color=self.RISK_COLORS[row['risk_level']],
                    end_color=self.RISK_COLORS[row['risk_level']],
                    fill_type='solid'
                )
    
    def _create_infrastructure_sheet(self, ws, df, include_charts):
        """Create infrastructure analysis sheet"""
        ws['A1'] = 'Infrastructure Access Analysis'
        ws['A1'].font = Font(bold=True, size=12)
        
        # Distance metrics
        ws['A3'] = 'Distance to Nearest Facility (km)'
        ws['A3'].font = Font(bold=True)
        
        distance_cols = [
            'dist_nearest_hospitals_km',
            'dist_nearest_fire_stations_km',
            'dist_nearest_ems_stations_km',
            'dist_nearest_nursing_homes_km'
        ]
        
        ws['A5'] = 'Facility Type'
        ws['B5'] = 'Mean'
        ws['C5'] = 'Median'
        ws['D5'] = 'Min'
        ws['E5'] = 'Max'
        ws['F5'] = 'Count > 50km'
        
        for idx, col in enumerate(distance_cols, start=6):
            if col in df.columns:
                facility_name = col.replace('dist_nearest_', '').replace('_km', '').replace('_', ' ').title()
                ws[f'A{idx}'] = facility_name
                ws[f'B{idx}'] = df[col].mean()
                ws[f'B{idx}'].number_format = '0.00'
                ws[f'C{idx}'] = df[col].median()
                ws[f'C{idx}'].number_format = '0.00'
                ws[f'D{idx}'] = df[col].min()
                ws[f'D{idx}'].number_format = '0.00'
                ws[f'E{idx}'] = df[col].max()
                ws[f'E{idx}'].number_format = '0.00'
                ws[f'F{idx}'] = (df[col] > 50).sum()
        
        # Facility counts
        ws['A13'] = 'Facilities Within 50km'
        ws['A13'].font = Font(bold=True)
        
        count_cols = [
            'count_hospitals_50km',
            'count_fire_stations_50km',
            'count_ems_stations_50km',
            'count_nursing_homes_50km'
        ]
        
        ws['A15'] = 'Facility Type'
        ws['B15'] = 'Mean'
        ws['C15'] = 'Median'
        ws['D15'] = 'Zero Count'
        
        for idx, col in enumerate(count_cols, start=16):
            if col in df.columns:
                facility_name = col.replace('count_', '').replace('_50km', '').replace('_', ' ').title()
                ws[f'A{idx}'] = facility_name
                ws[f'B{idx}'] = df[col].mean()
                ws[f'B{idx}'].number_format = '0.00'
                ws[f'C{idx}'] = df[col].median()
                ws[f'C{idx}'].number_format = '0.00'
                ws[f'D{idx}'] = (df[col] == 0).sum()
    
    def _create_demographics_sheet(self, ws, df, include_charts):
        """Create demographics analysis sheet"""
        ws['A1'] = 'Demographic Analysis'
        ws['A1'].font = Font(bold=True, size=12)
        
        # Population statistics
        ws['A3'] = 'Population Statistics'
        ws['A3'].font = Font(bold=True)
        
        pop_stats = [
            ('Total Population', df['total_population'].sum()),
            ('Mean County Population', df['total_population'].mean()),
            ('Median County Population', df['total_population'].median()),
            ('Largest County', df.loc[df['total_population'].idxmax(), 'county_name']),
            ('Smallest County', df.loc[df['total_population'].idxmin(), 'county_name'])
        ]
        
        for idx, (label, value) in enumerate(pop_stats, start=5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            if isinstance(value, (int, float)) and 'Population' in label:
                ws[f'B{idx}'].number_format = '#,##0'
        
        # Vulnerable populations
        ws['A12'] = 'Vulnerable Population Percentages'
        ws['A12'].font = Font(bold=True)
        
        vuln_cols = ['poverty_pct', 'elderly_pct', 'disability_pct', 'uninsured_pct']
        
        ws['A14'] = 'Indicator'
        ws['B14'] = 'Mean %'
        ws['C14'] = 'Median %'
        ws['D14'] = 'Max %'
        ws['E14'] = 'Counties > 25%'
        
        for idx, col in enumerate(vuln_cols, start=15):
            if col in df.columns:
                indicator_name = col.replace('_pct', '').replace('_', ' ').title()
                ws[f'A{idx}'] = indicator_name
                ws[f'B{idx}'] = df[col].mean()
                ws[f'B{idx}'].number_format = '0.00'
                ws[f'C{idx}'] = df[col].median()
                ws[f'C{idx}'].number_format = '0.00'
                ws[f'D{idx}'] = df[col].max()
                ws[f'D{idx}'].number_format = '0.00'
                ws[f'E{idx}'] = (df[col] > 25).sum()
    
    def _create_disaster_sheet(self, ws, df, include_charts):
        """Create disaster history sheet"""
        ws['A1'] = 'Disaster History Analysis'
        ws['A1'].font = Font(bold=True, size=12)
        
        # Overall disaster statistics
        ws['A3'] = 'Disaster Statistics'
        ws['A3'].font = Font(bold=True)
        
        disaster_stats = [
            ('Total Disasters (All Time)', df['disaster_count'].sum()),
            ('Mean Disasters per County', df['disaster_count'].mean()),
            ('Counties with No Disasters', (df['disaster_count'] == 0).sum()),
            ('Counties with 10+ Disasters', (df['disaster_count'] >= 10).sum())
        ]
        
        if 'disaster_count_recent' in df.columns:
            disaster_stats.extend([
                ('Recent Disasters (2015+)', df['disaster_count_recent'].sum()),
                ('Mean Recent per County', df['disaster_count_recent'].mean())
            ])
        
        for idx, (label, value) in enumerate(disaster_stats, start=5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            if isinstance(value, float):
                ws[f'B{idx}'].number_format = '0.00'
        
        # Disaster types
        disaster_type_cols = ['disaster_flood', 'disaster_hurricane', 'disaster_fire', 'disaster_tornado']
        
        if any(col in df.columns for col in disaster_type_cols):
            ws['A14'] = 'Disasters by Type'
            ws['A14'].font = Font(bold=True)
            
            ws['A16'] = 'Disaster Type'
            ws['B16'] = 'Total Count'
            ws['C16'] = 'Counties Affected'
            
            for idx, col in enumerate(disaster_type_cols, start=17):
                if col in df.columns:
                    disaster_name = col.replace('disaster_', '').title()
                    ws[f'A{idx}'] = disaster_name
                    ws[f'B{idx}'] = df[col].sum()
                    ws[f'C{idx}'] = (df[col] > 0).sum()
    
    def _create_high_risk_sheet(self, ws, df, include_charts):
        """Create high-risk counties sheet"""
        high_risk = df[df['risk_score'] >= 0.7].sort_values('risk_score', ascending=False)
        
        ws['A1'] = f'High Risk Counties (Score >= 0.7) - {len(high_risk)} counties'
        ws['A1'].font = Font(bold=True, size=12)
        
        if len(high_risk) == 0:
            ws['A3'] = 'No high-risk counties found'
            return
        
        # Summary statistics
        ws['A3'] = 'Summary Statistics'
        ws['A3'].font = Font(bold=True)
        
        stats = [
            ('Total High Risk Counties', len(high_risk)),
            ('Percentage of All Counties', f"{len(high_risk) / len(df) * 100:.1f}%"),
            ('Average Risk Score', high_risk['risk_score'].mean()),
            ('Highest Risk Score', high_risk['risk_score'].max()),
            ('Total Population at Risk', high_risk['total_population'].sum())
        ]
        
        for idx, (label, value) in enumerate(stats, start=5):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            if isinstance(value, float):
                ws[f'B{idx}'].number_format = '0.000'
        
        # List of high-risk counties
        ws['A12'] = 'High Risk County List'
        ws['A12'].font = Font(bold=True)
        
        display_cols = ['county_name', 'risk_score', 'risk_level', 'total_population', 'vulnerability_index']
        available_cols = [c for c in display_cols if c in high_risk.columns]
        
        # Headers
        for col_idx, col_name in enumerate(available_cols, 1):
            cell = ws.cell(row=14, column=col_idx, value=col_name.replace('_', ' ').title())
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Data
        for row_idx, (_, row) in enumerate(high_risk.iterrows(), 15):
            for col_idx, col_name in enumerate(available_cols, 1):
                value = row.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if pd.isna(value):
                    cell.value = ''
                else:
                    cell.value = value
                
                if col_name == 'risk_score' or col_name.endswith('_index'):
                    cell.number_format = '0.000'
                elif col_name == 'total_population':
                    cell.number_format = '#,##0'
    
    def _create_state_summary_sheet(self, ws, df, include_charts):
        """Create state summary sheet"""
        # Extract state from county_name
        df['state'] = df['county_name'].str.extract(r', ([\w ]+)$')[0]
        
        state_summary = df.groupby('state').agg({
            'fips': 'count',
            'total_population': 'sum',
            'risk_score': ['mean', 'max'],
            'vulnerability_index': 'mean',
            'isolation_index': 'mean',
            'disaster_count': 'sum'
        }).reset_index()
        
        state_summary.columns = ['State', 'Counties', 'Total Population', 
                                  'Avg Risk Score', 'Max Risk Score',
                                  'Avg Vulnerability', 'Avg Isolation', 'Total Disasters']
        
        ws['A1'] = 'State Summary'
        ws['A1'].font = Font(bold=True, size=12)
        
        # Headers
        headers = list(state_summary.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        # Data
        for row_idx, (_, row) in enumerate(state_summary.iterrows(), 4):
            for col_idx, col_name in enumerate(headers, 1):
                value = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                if col_name in ['Avg Risk Score', 'Max Risk Score', 'Avg Vulnerability', 'Avg Isolation']:
                    cell.number_format = '0.000'
                elif col_name == 'Total Population':
                    cell.number_format = '#,##0'
                elif col_name in ['Counties', 'Total Disasters']:
                    cell.number_format = '#,##0'
        
        # Sort by risk score
        state_summary_sorted = state_summary.sort_values('Avg Risk Score', ascending=False)
        
        ws['A' + str(len(state_summary) + 6)] = 'States by Risk (Highest First)'
        ws['A' + str(len(state_summary) + 6)].font = Font(bold=True)
        
        for idx, (_, row) in enumerate(state_summary_sorted.head(10).iterrows(), 
                                       start=len(state_summary) + 8):
            ws[f'A{idx}'] = row['State']
            ws[f'B{idx}'] = row['Avg Risk Score']
            ws[f'B{idx}'].number_format = '0.000'


def main():
    """CLI for Excel export"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Export ResilienceAI data as Excel")
    parser.add_argument("--sheets", nargs="+", 
                        default=["Summary", "County Data", "Risk Analysis"],
                        help="Sheets to include")
    parser.add_argument("--output", "-o", help="Output file path")
    parser.add_argument("--state", help="Filter by state")
    parser.add_argument("--risk-level", choices=["Low", "Medium", "High"],
                        help="Filter by risk level")
    parser.add_argument("--min-risk", type=float, help="Minimum risk score")
    parser.add_argument("--no-charts", action="store_true", help="Exclude charts")
    
    args = parser.parse_args()
    
    exporter = ExcelExporter()
    
    if exporter.df is None:
        print("Error: County data not found. Run pipeline first.")
        return
    
    filters = {}
    if args.state:
        filters['state'] = args.state
    if args.risk_level:
        filters['risk_level'] = args.risk_level
    if args.min_risk:
        filters['min_risk_score'] = args.min_risk
    
    result = exporter.export_workbook(
        sheets=args.sheets,
        output_path=args.output,
        filters=filters if filters else None,
        include_charts=not args.no_charts
    )
    
    if "error" in result:
        print(f"Error: {result['error']}")
    else:
        print(f"Exported to: {result['output_path']}")
        print(f"Sheets: {', '.join(result['sheets'])}")
        print(f"Row count: {result['row_count']}")


if __name__ == "__main__":
    main()
